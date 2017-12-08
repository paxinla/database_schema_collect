#!/usr/bin/env python
# coding=utf-8

from __future__ import absolute_import
from __future__ import division

import sys
import os
import logging
import traceback

from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors, PatternFill, Border, Side, Alignment

reload(sys)
sys.setdefaultencoding("utf-8")


logger = logging.getLogger("database_schema_collect")


class DataDictionary(object):
    """Generate data dictionary to Excel file."""

    def __init__(self, dbname, bank_obj, work_dir):
        self.bank_obj = bank_obj
        self.work_dir = work_dir
        self.dbname = dbname

        self._excel_name = "{}.xlsx".format(os.path.join(self.work_dir, self.dbname))
        self.des_excel = Workbook()
        self.sheet_index = 0

        self.default_font = Font(name="Arial", size=12)
        self.default_link_font = Font(name="Arial", size=12, color=colors.BLUE)
        self.default_header_font = Font(name="Arial", size=12, bold=True, color=colors.WHITE)
        self.default_header_align = Alignment(horizontal="center", vertical="center", shrink_to_fit=False)
        self.default_header_border = Border(left=Side(border_style="thick",
                                                      color='FF000000'),
                                            right=Side(border_style="thick",
                                                       color='FF000000'),
                                            top=Side(border_style="thick",
                                                     color='FF000000'),
                                            bottom=Side(border_style="thick",
                                                        color='FF000000'))
        self.default_border = Border(left=Side(border_style="thin",
                                               color='FF000000'),
                                     right=Side(border_style="thin",
                                                color='FF000000'),
                                     top=Side(border_style="thin",
                                              color='FF000000'),
                                     bottom=Side(border_style="thin",
                                                 color='FF000000'))
        self.default_align = Alignment(shrink_to_fit=False)


    def save_excel(self):
        """Save the excel work book."""
        self.des_excel.save(self._excel_name)


    def _change_cell_style(self, work_sheet, cell_range, font=None, border=None, fill=None, align=None):
        """Apply style on cells.

        :param work_sheet: Target work sheet.
        :type work_sheet: An instance of work sheet.
        :param cell_range: Target range of cells.
        :type cell_range: str, format like "A3:D5".
        :param font: Style of font.
        :type font: An instance of Font.
        :param border: Style of border.
        :type border: An instance of border.
        :param fill: Style of cell fill up.
        :type fill: An instance of Fill.
        :param align: Style of alignment.
        :type align: An instance of alignment.
        """
        try:
            rows = work_sheet[cell_range]

            for row in rows:
                for c in row:
                    if font:
                        c.font = font
                    else:
                        c.font = self.default_font

                    if border:
                        c.border = border
                    else:
                        c.border = self.default_border

                    if fill:
                        c.fill = fill

                    if align:
                        c.alignment = align
                    else:
                        c.alignment = self.default_align
        except:
            logger.error(traceback.format_exc())


    def gen_toc_sheet(self):
        """Generate a list of all tables and views."""
        logger.info("Generate table of contents.")

        ws_toc = self.des_excel.create_sheet("Table of contents", self.sheet_index)
        self.sheet_index += 1
        ws_toc.sheet_properties.tabColor = "FCA405"

        ws_toc["A1"] = "List of sheets"
        ws_toc["A1"].font = self.default_font


        ws_toc["A3"] = "Schema"
        ws_toc["B3"] = "Name"
        ws_toc["C3"] = "Type"
        ws_toc["D3"] = "Description"
        self._change_cell_style(ws_toc,
                                "A3:D3",
                                self.default_header_font,
                                self.default_header_border,
                                PatternFill("solid", fgColor="FCA405"),
                                self.default_header_align)

        # Link to tablespaces
        ws_toc["A4"] = '-'
        ws_toc["B4"] = '=HYPERLINK("{}", "{}")'.format("#Tablespaces!A2", "Tablespaces")
        ws_toc["C4"] = "tablespace"
        ws_toc["D4"] = "List of tablespaces."
        self._change_cell_style(ws_toc, "A4:D4", None, None, None, None)
        ws_toc["B4"].font = self.default_link_font

        ridx_start = 5
        ridx = ridx_start - 1
        # Link to all tables
        if self.bank_obj.tables is not None and self.bank_obj.tables != []:
            sorted_tables = sorted(self.bank_obj.tables, key=lambda x: x.table_schemaname)
            for each_table in sorted_tables:
                ridx += 1
                fulltablename = "{}.{}".format(each_table.table_schemaname, each_table.table_name)
                ws_toc.cell(row=ridx, column=1).value = each_table.table_schemaname
                ws_toc.cell(row=ridx, column=2).value = '=HYPERLINK("{}", "{}")'.format("#{}!A2".format(fulltablename), fulltablename)
                ws_toc.cell(row=ridx, column=3).value = "table"
                ws_toc.cell(row=ridx, column=4).value = each_table.table_comment

        # Link to all views
        if self.bank_obj.views is not None and self.bank_obj.views != []:
            sorted_views = sorted(self.bank_obj.views, key=lambda x: x.view_name)
            for each_view in sorted_views:
                ridx += 1
                fullviewname = "{}.{}".format(each_view.view_schemaname, each_view.view_name)
                ws_toc.cell(row=ridx, column=1).value = each_view.view_schemaname
                ws_toc.cell(row=ridx, column=2).value = '=HYPERLINK("{}", "{}")'.format("#{}!A2".format(fullviewname), fullviewname)
                ws_toc.cell(row=ridx, column=3).value = "view"
                ws_toc.cell(row=ridx, column=4).value = "-"

        cell_range = "A{!s}:D{!s}".format(ridx_start, ridx)
        logger.debug("[TOC] Apply style to range: %s" % cell_range)

        self._change_cell_style(ws_toc, cell_range, None, None, None, None)
        self._change_cell_style(ws_toc,
                                "B{!s}:B{!s}".format(ridx_start, ridx),
                                self.default_link_font,
                                None, None, None)


    def gen_tablspaces_sheet(self):
        """Generate a list of all tablespaces."""
        logger.info("Generate one sheet for list tablespaces.")

        ws_tbs = self.des_excel.create_sheet("Tablespaces", self.sheet_index)
        self.sheet_index += 1
        ws_tbs.sheet_properties.tabColor = "14CAD4"

        ws_tbs["A1"] = "List of tablespaces"
        ws_tbs["A1"].font = self.default_font

        ws_tbs["A3"] = "Name"
        ws_tbs["B3"] = "Location"
        ws_tbs["C3"] = "Description"
        ws_tbs["D3"] = "Owner"
        self._change_cell_style(ws_tbs,
                                "A3:D3",
                                self.default_header_font,
                                self.default_header_border,
                                PatternFill("solid", fgColor="14CAD4"),
                                self.default_header_align)

        ridx_start = 3
        ridx = ridx_start
        for each_tbs in self.bank_obj.tablespaces:
            ridx += 1
            ws_tbs.cell(row=ridx, column=1).value = each_tbs.tablespace_name
            ws_tbs.cell(row=ridx, column=2).value = each_tbs.tablespace_location
            ws_tbs.cell(row=ridx, column=3).value = each_tbs.tablespace_comment
            ws_tbs.cell(row=ridx, column=4).value = each_tbs.tablespace_owner
        cell_range = "A{!s}:D{!s}".format(ridx_start+1, ridx)
        logger.debug("[Tablespaces] Apply style to range: %s" % cell_range)
        self._change_cell_style(ws_tbs, cell_range, None, None, None, None)


    def gen_table_sheet(self):
        """Generate a sheet contains information of a table."""
        logger.info("Generate one sheet for one table, including indexes, foreign keys and other constraints belong to this table.")

        if self.bank_obj.tables is None or self.bank_obj.tables == []:
            return

        ptn_table_header = PatternFill("solid", fgColor="191995")
        ptn_column_header = PatternFill("solid", fgColor="14CAD4")
        ptn_pk_header = PatternFill("solid", fgColor="E3C381")
        ptn_uk_header = PatternFill("solid", fgColor="C1BA7E")
        ptn_fk_header = PatternFill("solid", fgColor="E8A782")
        ptn_ck_header = PatternFill("solid", fgColor="D4C69B")
        ptn_idx_header = PatternFill("solid", fgColor="EE82EE")

        sorted_tables = sorted(self.bank_obj.tables, key=lambda x: x.table_schemaname)
        for each_table in sorted_tables:
            fulltablename = "{}.{}".format(each_table.table_schemaname, each_table.table_name)
            logger.info("Create sheet for table %s" % fulltablename)
            ws_tb = self.des_excel.create_sheet(fulltablename, self.sheet_index)
            self.sheet_index += 1

            # Tablename part
            ws_tb["A1"] = "TableName: "
            ws_tb["B1"] = each_table.table_name
            ws_tb["A2"] = "Description"
            ws_tb["B2"] = each_table.table_comment
            ws_tb["A3"] = "Owner"
            ws_tb["B3"] = each_table.table_owner

            self._change_cell_style(ws_tb,
                                    "A1:A3",
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_table_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb, "B1:B3", None, None, None, None)

            # Column part
            ws_tb["A5"] = "Column Name"
            ws_tb["B5"] = "Column Type"
            ws_tb["C5"] = "Length"
            ws_tb["D5"] = "Scale"
            ws_tb["E5"] = "Precision"
            ws_tb["F5"] = "Not Null?"
            ws_tb["G5"] = "Auto Inc?"
            ws_tb["H5"] = "Default"
            ws_tb["I5"] = "Description"

            self._change_cell_style(ws_tb,
                                    "A5:I5",
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_column_header,
                                    self.default_header_align)

            sorted_columns = sorted(each_table.columns, key=lambda x: x.column_index)
            ridx_start = 6
            ridx = ridx_start - 1
            for each_column in sorted_columns:
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_column.column_name
                ws_tb.cell(row=ridx, column=2).value = each_column.column_data_type
                ws_tb.cell(row=ridx, column=3).value = each_column.column_length
                ws_tb.cell(row=ridx, column=4).value = each_column.column_scale
                ws_tb.cell(row=ridx, column=5).value = each_column.column_precision
                ws_tb.cell(row=ridx, column=6).value = 'N' if each_column.column_is_nullable else 'Y'
                ws_tb.cell(row=ridx, column=7).value = each_column.column_auto_increment
                ws_tb.cell(row=ridx, column=8).value = each_column.column_default_value
                ws_tb.cell(row=ridx, column=9).value = each_column.column_comment
            col_cell_range = "A{!s}:I{!s}".format(ridx_start, ridx)
            logger.debug("[columns] Apply style to range: %s" % col_cell_range)
            self._change_cell_style(ws_tb, col_cell_range, None, None, None, None)

            # Primary key part
            ridx += 1
            ridx_start = ridx + 1
            ridx = ridx_start

            ws_tb.cell(row=ridx, column=1).value = "Primary Key"
            ridx += 1
            ws_tb.cell(row=ridx, column=1).value = "Key Name"
            ws_tb.cell(row=ridx, column=2).value = "Column Name"
            ws_tb.cell(row=ridx, column=3).value = "Column Index"
            ws_tb.cell(row=ridx, column=4).value = "Index Tablespace"
            self._change_cell_style(ws_tb,
                                    "A{i!s}:A{i!s}".format(i=ridx_start),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_pk_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb,
                                    "A{!s}:D{!s}".format(ridx_start+1, ridx),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_pk_header,
                                    self.default_header_align)

            for each_pk in sorted(each_table.primary_key, key=lambda x:x.pk_column_index):
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_pk.pk_name
                ws_tb.cell(row=ridx, column=2).value = each_pk.pk_column
                ws_tb.cell(row=ridx, column=3).value = each_pk.pk_column_index
                ws_tb.cell(row=ridx, column=4).value = each_pk.pk_tablespace
            pk_cell_range = "A{!s}:D{!s}".format(ridx_start+2, ridx)
            logger.debug("[primary key] Apply style to range %s" % pk_cell_range)
            self._change_cell_style(ws_tb, pk_cell_range, None, None, None, None)

            # Unique key part
            ridx += 1
            ridx_start = ridx + 1
            ridx = ridx_start

            ws_tb.cell(row=ridx, column=1).value = "Unique Keys"
            ridx += 1
            ws_tb.cell(row=ridx, column=1).value = "Key Name"
            ws_tb.cell(row=ridx, column=2).value = "Column Name"
            ws_tb.cell(row=ridx, column=3).value = "Column Index"
            ws_tb.cell(row=ridx, column=4).value = "Index Tablespace"
            self._change_cell_style(ws_tb,
                                    "A{i!s}:A{i!s}".format(i=ridx_start),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_uk_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb,
                                    "A{!s}:D{!s}".format(ridx_start+1, ridx),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_uk_header,
                                    self.default_header_align)

            for each_uk in sorted(each_table.unique_keys, key=lambda x: (x.uk_name, x.uk_column_index)):
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_uk.uk_name
                ws_tb.cell(row=ridx, column=2).value = each_uk.uk_column
                ws_tb.cell(row=ridx, column=3).value = each_uk.uk_column_index
                ws_tb.cell(row=ridx, column=4).value = each_uk.uk_tablespace
            uk_cell_range = "A{!s}:D{!s}".format(ridx_start+2, ridx)
            logger.debug("[unique key] Apply style to range %s" % uk_cell_range)
            self._change_cell_style(ws_tb, uk_cell_range, None, None, None, None)

            # Check condition part
            ridx += 1
            ridx_start = ridx + 1
            ridx = ridx_start

            ws_tb.cell(row=ridx, column=1).value = "Check Constraints"
            ridx += 1
            ws_tb.cell(row=ridx, column=1).value = "Check Name"
            ws_tb.cell(row=ridx, column=2).value = "Constraint define"
            self._change_cell_style(ws_tb,
                                    "A{i!s}:A{i!s}".format(i=ridx_start),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_ck_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb,
                                    "A{!s}:B{!s}".format(ridx_start+1, ridx),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_ck_header,
                                    self.default_header_align)

            for each_ck in sorted(each_table.checks, key=lambda x: x.check_name):
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_ck.check_name
                ws_tb.cell(row=ridx, column=2).value = each_ck.check_define
            ck_cell_range = "A{!s}:B{!s}".format(ridx_start+2, ridx)
            logger.debug("[check] Apply style to range %s" % ck_cell_range)
            self._change_cell_style(ws_tb, ck_cell_range, None, None, None, None)

            # Foreign key part
            ridx += 1
            ridx_start = ridx + 1
            ridx = ridx_start

            ws_tb.cell(row=ridx, column=1).value = "Foreign Keys"
            ridx += 1
            ws_tb.cell(row=ridx, column=1).value = "Key Name"
            ws_tb.cell(row=ridx, column=2).value = "Column Name"
            ws_tb.cell(row=ridx, column=3).value = "Reference Table Name"
            ws_tb.cell(row=ridx, column=4).value = "Reference Column Name"
            self._change_cell_style(ws_tb,
                                    "A{i!s}:A{i!s}".format(i=ridx_start),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_fk_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb,
                                    "A{!s}:D{!s}".format(ridx_start+1, ridx),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_fk_header,
                                    self.default_header_align)

            for each_fk in sorted(each_table.foreign_keys, key=lambda x: x.fk_name):
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_fk.fk_name
                ws_tb.cell(row=ridx, column=2).value = each_fk.fk_column
                ws_tb.cell(row=ridx, column=3).value = each_fk.fk_ref_tablename
                ws_tb.cell(row=ridx, column=4).value = each_fk.fk_ref_column
            fk_cell_range = "A{!s}:D{!s}".format(ridx_start+2, ridx)
            logger.debug("[foreign key] Apply style to range %s" % fk_cell_range)
            self._change_cell_style(ws_tb, fk_cell_range, None, None, None, None)

            # Index part
            ridx += 1
            ridx_start = ridx + 1
            ridx = ridx_start

            ws_tb.cell(row=ridx, column=1).value = "Indexes"
            ridx += 1
            ws_tb.cell(row=ridx, column=1).value = "Index Name"
            ws_tb.cell(row=ridx, column=2).value = "Index Type"
            ws_tb.cell(row=ridx, column=4).value = "Columns"
            ws_tb.cell(row=ridx, column=3).value = "Tablespace"
            self._change_cell_style(ws_tb,
                                    "A{i!s}:A{i!s}".format(i=ridx_start),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_idx_header,
                                    self.default_header_align)
            self._change_cell_style(ws_tb,
                                    "A{!s}:D{!s}".format(ridx_start+1, ridx),
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_idx_header,
                                    self.default_header_align)

            for each_idx in sorted(each_table.indexes, key=lambda x: x.index_name):
                ridx += 1
                ws_tb.cell(row=ridx, column=1).value = each_idx.index_name
                ws_tb.cell(row=ridx, column=2).value = each_idx.index_type
                ws_tb.cell(row=ridx, column=3).value = each_idx.index_columns
                ws_tb.cell(row=ridx, column=4).value = each_idx.index_tablespace
            idx_cell_range = "A{!s}:D{!s}".format(ridx_start+2, ridx)
            logger.debug("[index] Apply style to range %s" % idx_cell_range)
            self._change_cell_style(ws_tb, idx_cell_range, None, None, None, None)


    def gen_view_sheet(self):
        """Generate a sheet contains information of a view."""
        logger.info("Generate one sheet for one view.")

        if self.bank_obj.views is None or self.bank_obj.views == []:
            return

        ptn_view_header = PatternFill("solid", fgColor="191995")
        ptn_sql_header = PatternFill("solid", fgColor="14CAD4")

        for each_view in self.bank_obj.views:
            fullviewname = "{}.{}".format(each_view.view_schemaname, each_view.view_name)
            logger.info("Create sheet for view %s" % fullviewname)
            ws_vw = self.des_excel.create_sheet(fullviewname, self.sheet_index)
            self.sheet_index += 1

            # Viewname part
            ws_vw["A1"] = "View Name: "
            ws_vw["B1"] = each_view.view_name
            ws_vw["A2"] = "Description"
            ws_vw["B2"] = each_view.view_comment
            ws_vw["A3"] = "Owner"
            ws_vw["B3"] = each_view.view_owner

            self._change_cell_style(ws_vw,
                                    "A1:A3",
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_view_header,
                                    self.default_header_align)
            self._change_cell_style(ws_vw, "B1:B3", None, None, None, None)

            # Define SQL
            ws_vw["A5"] = "Define SQL"
            self._change_cell_style(ws_vw,
                                    "A5:A5",
                                    self.default_header_font,
                                    self.default_header_border,
                                    ptn_sql_header,
                                    self.default_header_align)
            ws_vw["A6"] = each_view.view_define
            self._change_cell_style(ws_vw, "A6:A6", None, None, None,
                                    Alignment(wrap_text=True,
                                              shrink_to_fit=True,))


    def gen_dictionanry(self):
        """Generate the whole data dictionanry"""
        self.gen_toc_sheet()
        self.gen_tablspaces_sheet()
        self.gen_table_sheet()
        self.gen_view_sheet()
        self.save_excel()
